import openpyxl

def get_rowcount(file,sheet):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet]
    return sheet.max_row

def get_colcount(file,sheet):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet]
    return sheet.max_column

def read_data(file,sheet,r_num,c_num):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet]
    return sheet.cell(row=r_num,column=c_num).value

def write_data(file,sheet,r_num,c_num,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet]
    sheet.cell(row=r_num,column=c_num).value=data
    workbook.save(file)